from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_tables(filename):
    wb = load_workbook(filename)
    style = TableStyleInfo(
        name='TableStyleMedium5',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if len(list(ws.values)) == 0:
            continue

        table_name = sheet.replace(' ', '_')
        tab = Table(displayName=table_name, ref=ws.dimensions)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        print(f"{sheet} table generated")

    wb.save(filename)
