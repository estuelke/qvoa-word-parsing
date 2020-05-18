"""
Transfers all tables from the master word document to Excel Files that will be
processed by the application.  Not used as part of the main application.
"""
import pandas as pd
import re
from docx import Document
from docx.oxml.text.paragraph import CT_P
from openpyxl import Workbook

masterfilepath = 'files/qvoa_master_file.docx'
masterfile = Document(masterfilepath)
ph_num_patterns = [
    r'[Pp][Hh]\s*?[#:]*?\s*?',
    r'(#\s*?86\s*?)*#\s*?',
    r'Pheresis\s*?'
]
ph_num_pattern = rf"^\s*?({('|').join(ph_num_patterns)})(?P<ph_num>\d{{1,3}})[\sMD_]"

patterns = [
    r'g?Millions?\/\s*?m[Ll]',
    r'cor',
    r'Dilution\s+\([Mm]illion\)',
    r'[15]',
    r'IUPMs?',
    r'Day\s+\d+\s+in\s+[Cc]ulture?(\s*\(GSK\))?',
    r'Day 19 \(PHA and reg IL-2\) or Day 20 \(GSK, SAHA and CTRLS\) in culture',
    # r'Day 19 in culture .*?',
    r'D\d+(\/\d+)?\s+in\s+culture.*?',
    r'D\d+',
    r'',
    r'\s*?Day\s+\d+\/\d+\s+in Culture',
    r'D19\s+Millions\/ml',
    r'2\.5\s+\(1\.0\)',
    r'[0123]\.[01256][01245]?[25]?[5]?',
    r'\*\s+Mold'
]
first_column_pattern = rf"^\s*?({('|').join(patterns)})\s*?$"
file_counter = 1
sheet_counter = 1
wb = Workbook()
tables = []


def get_num(row_text):
    for text in row_text:
        ph_num = re.match(ph_num_pattern, text)
        if ph_num:
            return ph_num.group('ph_num')
    return None


for table in masterfile.tables:
    table_info = {}
    for excel_row, word_row in enumerate(table.rows, 1):
        row_text = [cell.text for cell in word_row.cells]

        if excel_row == 1:
            if re.match(r'.new. UNC method|Old UTSW', row_text[0]):
                table_info['odd_top_row'] = list(set(row_text))
                next_row = table.rows[excel_row]
                row_text = [cell.text for cell in next_row.cells]
            row_text = list(set(row_text))
            ph_num = get_num(row_text)

            if ph_num:
                ws = wb.create_sheet(ph_num)
                table_info['ph_num'] = ph_num
            else:
                ws = wb.create_sheet()

            table_info['first_row'] = row_text
            table_info['ws_title'] = ws.title
            fix_table = []

        for excel_col, cell_text in enumerate(row_text, 1):
            if excel_col == 1 and excel_row > 1 and not re.match(
                first_column_pattern, cell_text
            ):
                fix_table.append(cell_text)

            ws.cell(column=excel_col, row=excel_row).value = cell_text
    table_info['fixes'] = fix_table
    table_info['file_number'] = file_counter
    tables.append(table_info)

    if 'odd_top_row' in table_info:
        row_count = len(table.rows) + 2
        for text in table_info['odd_top_row']:
            ws.cell(column=1, row=row_count).value = text
            row_count += 1

    row_count = len(table.rows) + 2
    next_element = table._tbl.getnext()
    while isinstance(next_element, CT_P):
        text_box = next_element.xpath('.//w:txbxContent//w:t')
        if text_box:
            for item in text_box:
                ws.cell(column=1, row=row_count).value = \
                    f"TABLE_NOTE: {item.text}"
                row_count += 1
        next_element = next_element.getnext()

    sheet_counter += 1
    if sheet_counter // 21 == 1:
        wb.save(f"files/QVOA_Tables_{file_counter:02}.xlsx")
        file_counter += 1
        wb = Workbook()
        sheet_counter = 1

df = pd.DataFrame(tables)
df.to_excel('Errors to Manually Fix.xlsx')
