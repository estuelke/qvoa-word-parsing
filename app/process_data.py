import os
import pandas as pd
from openpyxl import load_workbook
from app.clean.clean import distribute_and_clean_data
from app.helpers import names
from app.sheet import process_sheet


def process_file(wb, filename):
    for sheetname in wb.sheetnames:
        if sheetname == names.SHEET:
            continue
        ws = wb[sheetname]
        yield from process_sheet(ws, filename)


def process_files(filenames):

    for filename in filenames:
        wb = load_workbook(filename)

        head, tail = os.path.split(filename)
        fn, ext = os.path.splitext(tail)

        yield from process_file(wb, fn)

    # Used as reference point when viewing in Jupyter
    print(f"All QVOA Tables files processed.")


def process_data(filenames):
    data = process_files(filenames)
    data = pd.DataFrame(data)
    data = distribute_and_clean_data(data)

    return data
